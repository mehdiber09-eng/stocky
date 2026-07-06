"""Export utilities for lots and stock reports."""
from fastapi import APIRouter, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from datetime import datetime, timezone, timedelta
from typing import Optional
import io
import json

from app.api.deps import get_db, get_current_user
from app.models import models
from app.core.config import settings

router = APIRouter()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


@router.get("/export/expiring-lots-xlsx")
async def export_expiring_lots_xlsx(
    days: int = Query(30),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Export expiring lots as Excel file."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}

    now = datetime.now(timezone.utc)
    cutoff = now + timedelta(days=days)

    stmt = select(models.Lot).where(
        models.Lot.owner_id == user.id,
        models.Lot.expiry_date != None,
        models.Lot.expiry_date <= cutoff,
    ).order_by(models.Lot.expiry_date)

    result = await db.execute(stmt)
    lots = result.scalars().all()

    # Fetch product names
    product_map = {}
    for lot in lots:
        if lot.product_id not in product_map:
            product = await db.get(models.Product, lot.product_id)
            product_map[lot.product_id] = product.name if product else f"Produit #{lot.product_id}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Lots à écouler"

    # Header
    headers = [
        "Produit",
        "Code lot",
        "Date d'expiration",
        "Jours restants",
        "Quantité totale",
        "Quantité disponible",
        "Référence fournisseur",
        "Date de réception",
        "Urgence",
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for lot in lots:
        if lot.expiry_date:
            expiry = lot.expiry_date.strftime("%Y-%m-%d")
            days_left = (lot.expiry_date - now).days
            urgency = "🔴 CRITIQUE" if days_left <= 7 else "🟠 ATTENTION" if days_left <= 14 else "🟡 À SUIVRE"
        else:
            expiry = ""
            days_left = ""
            urgency = ""

        product_name = product_map.get(lot.product_id, f"Produit #{lot.product_id}")

        row = [
            product_name,
            lot.batch_code or "",
            expiry,
            days_left,
            lot.quantity_total,
            lot.quantity_available,
            lot.supplier_lot_ref or "",
            lot.received_at.strftime("%Y-%m-%d") if lot.received_at else "",
            urgency,
        ]
        ws.append(row)

        # Color code by urgency
        if days_left is not None:
            if days_left <= 3:
                fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                font = Font(color="FFFFFF")
            elif days_left <= 7:
                fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                font = Font(color="FFFFFF")
            elif days_left <= 14:
                fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                font = Font(color="000000")
            else:
                fill = None
                font = None

            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
                    cell.font = font

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15

    # Summary
    ws.append([])
    ws.append(["Résumé"])
    ws.append([f"Total lots: {len(lots)}"])
    ws.append([f"Critiques (≤7j): {sum(1 for lot in lots if lot.expiry_date and (lot.expiry_date - now).days <= 7)}"])
    ws.append([f"Généré le: {now.strftime('%Y-%m-%d %H:%M:%S')}"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return {
        "data": output.getvalue().hex(),
        "filename": f"lots-expirant-{now.strftime('%Y%m%d')}.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


@router.get("/export/expiring-lots-json")
async def export_expiring_lots_json(
    days: int = Query(30),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Export expiring lots as JSON (for client-side PDF generation)."""
    now = datetime.now(timezone.utc)
    cutoff = now + timedelta(days=days)

    stmt = select(models.Lot).where(
        models.Lot.owner_id == user.id,
        models.Lot.expiry_date != None,
        models.Lot.expiry_date <= cutoff,
    ).order_by(models.Lot.expiry_date)

    result = await db.execute(stmt)
    lots = result.scalars().all()

    # Fetch product info
    product_map = {}
    for lot in lots:
        if lot.product_id not in product_map:
            product = await db.get(models.Product, lot.product_id)
            if product:
                product_map[lot.product_id] = {
                    "name": product.name,
                    "sku": product.sku,
                    "unit_price": product.unit_price,
                    "cost_price": product.cost_price,
                }

    export_data = {
        "generated_at": now.isoformat(),
        "export_window_days": days,
        "total_lots": len(lots),
        "lots": [
            {
                "id": lot.id,
                "product_id": lot.product_id,
                "product_name": product_map.get(lot.product_id, {}).get("name", f"Produit #{lot.product_id}"),
                "batch_code": lot.batch_code,
                "expiry_date": lot.expiry_date.isoformat() if lot.expiry_date else None,
                "received_at": lot.received_at.isoformat() if lot.received_at else None,
                "quantity_total": lot.quantity_total,
                "quantity_available": lot.quantity_available,
                "supplier_lot_ref": lot.supplier_lot_ref,
                "days_until_expiry": (lot.expiry_date - now).days if lot.expiry_date else None,
                "urgency": (
                    "CRITICAL" if lot.expiry_date and (lot.expiry_date - now).days <= 3
                    else "WARNING" if lot.expiry_date and (lot.expiry_date - now).days <= 7
                    else "WATCH"
                ),
            }
            for lot in lots
        ],
        "summary": {
            "critical": sum(1 for lot in lots if lot.expiry_date and (lot.expiry_date - now).days <= 3),
            "warning": sum(1 for lot in lots if lot.expiry_date and (lot.expiry_date - now).days <= 7),
            "watch": sum(1 for lot in lots if lot.expiry_date and (lot.expiry_date - now).days > 7),
        },
    }

    return export_data
